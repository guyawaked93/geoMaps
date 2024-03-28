import folium
from folium.plugins import MarkerCluster, Geocoder
import openpyxl
from concurrent.futures import ThreadPoolExecutor

def create_marker(coord):
    geo_correta = coord.get('GeoCorreta')
    geo_correta_line = f"<b>GeoCorreta:</b> {geo_correta}<br>" if geo_correta else ""

    popup_info = f"""
        <b>{coord['Nome da Escola']}</b><br>
        <i>{coord['Município']}</i>, {coord['UF']}<br>
        <b>Endereço:</b> {coord['Endereço']}<br>
        <b>Kit Gerador Solar:</b> {coord.get('Kit Gerador Solar (estimado)', 'N/A')}<br>
        <b>Kit Instalação Elétrica Interna:</b> {coord.get('Kit Instalação Elétrica Interna (estimado)', 'N/A')}<br>
        {geo_correta_line}
    """
    return folium.Marker(location=[coord["Latitude"] + offset, coord["Longitude"] + offset], popup=folium.Popup(popup_info, max_width=300), icon=folium.Icon(color='green'))

def create_marker_red(coord):
    popup_info = f"""
        <b>{coord['Nome da Escola']}</b><br>
        <i>{coord['Município']}</i>, {coord['UF']}<br>
        <b>Endereço:</b> {coord['Endereço']}<br>
        <b>Kit Wi-Fi (estimado):</b> {coord.get('Kit Wi-Fi (estimado)', 'N/A')}<br>
        <b>AP adicional (estimado):</b> {coord.get('AP adicional (estimado)', 'N/A')}<br>
        <b>Nobreak:</b> {coord.get('Nobreak', 'N/A')}<br>
    """
    return folium.Marker(location=[coord["Latitude"], coord["Longitude"]], popup=folium.Popup(popup_info, max_width=300), icon=folium.Icon(color='red'))

offset = 0.0005

wb = openpyxl.load_workbook("escolas.xlsx")
sheet = wb.active

coordenadas = []
for row in sheet.iter_rows(min_row=2): 
    latitude = row[5].value
    longitude = row[6].value
    if latitude is not None and longitude is not None and isinstance(latitude, (int, float)) and isinstance(longitude, (int, float)):
        coordenadas.append({
            "UF": row[0].value,
            "Município": row[1].value,
            "Código INEP": row[2].value,
            "Nome da Escola": row[3].value,
            "Endereço": row[4].value,
            "Latitude": float(latitude),
            "Longitude": float(longitude),
            "Kit Gerador Solar (estimado)": row[7].value,
            "Kit Instalação Elétrica Interna (estimado)": row[8].value,
            "GeoCorreta": row[9].value
        })

wb_redimensionamentoz = openpyxl.load_workbook("redimensionamentoz.xlsx")
sheet_redimensionamentoz = wb_redimensionamentoz.active 

coordenadas_redimensionamentoz = []
for row in sheet_redimensionamentoz.iter_rows(min_row=2): 
    latitude = row[5].value
    longitude = row[6].value
    if latitude is not None and longitude is not None and isinstance(latitude, (int, float)) and isinstance(longitude, (int, float)):
        coordenadas_redimensionamentoz.append({
            "UF": row[0].value,
            "Município": row[1].value,
            "Código INEP": row[2].value,
            "Nome da Escola": row[3].value,
            "Endereço": row[4].value,
            "Latitude": float(latitude),
            "Longitude": float(longitude),
            "Kit Wi-Fi (estimado)": row[7].value,
            "AP adicional (estimado)": row[8].value,
            "Nobreak": row[9].value
        })

m = folium.Map(location=[-15.788, -47.879], zoom_start=4)

cluster_electricity = MarkerCluster(name='Electricidade').add_to(m)
cluster_wifi = MarkerCluster(name='Wifi').add_to(m)

with ThreadPoolExecutor() as executor:
    markers = list(executor.map(create_marker, coordenadas))

for marker in markers:
    marker.add_to(cluster_electricity)

with ThreadPoolExecutor() as executor:
    markers_red = list(executor.map(create_marker_red, coordenadas_redimensionamentoz))

for marker in markers_red:
    marker.add_to(cluster_wifi)

folium.LayerControl().add_to(m)
folium.plugins.LocateControl().add_to(m)

Geocoder().add_to(m)


sidebar_html = """
<div style="position: fixed; top: 50px; left: 50px; width: 200px; height: 400px; overflow: auto; padding: 5px; border:2px solid black; background-color: transparent; z-index:9999" id="sidebar">
    <h4>Escolas</h4>
    <ul>
"""

for coord in coordenadas:
    sidebar_html += f'<li><a href="#" onclick="m.setView([{coord["Latitude"]}, {coord["Longitude"]}]); return false;">{coord["Nome da Escola"]}</a></li>'

sidebar_html += """
    </ul>
</div>
<button style="position: fixed; top: 120px; left: 10px; z-index:9999" onclick="toggleSidebar()">☰</button>
<script>
    var sidebarVisible = true;
    function toggleSidebar() {
        var sidebar = document.getElementById('sidebar');
        if (sidebarVisible) {
            sidebar.style.display = 'none';
        } else {
            sidebar.style.display = 'block';
        }
        sidebarVisible = !sidebarVisible;
    }
</script>
"""

m.get_root().html.add_child(folium.Element(sidebar_html))

m.save("index.html")